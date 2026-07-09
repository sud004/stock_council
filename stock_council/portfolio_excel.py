#!/usr/bin/env python3
# ============================================================
# portfolio_excel.py — Portfolio Tracker Excel
# ============================================================
# Writes data/excel/portfolio_tracker.xlsx every night.
# 5 sheets:
#   1. Daily Log    — one row per position per day
#   2. Daily Summary — one row per day (P&L, cash, mode)
#   3. Open Positions — live snapshot
#   4. Closed Trades — completed trades
#   5. Learning Feedback — weight changes log
# ============================================================

import sys
import time
from pathlib import Path
from datetime import datetime
import pytz

sys.path.insert(0, str(Path(__file__).parent))

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
except ImportError:
    print("[PORTFOLIO EXCEL] Install: pip install openpyxl pandas")
    sys.exit(1)

from config import DATA_DIR
from portfolio_engine import (
    STARTING_CAPITAL, load_state, _portfolio_value
)

IST             = pytz.timezone('Asia/Kolkata')
PORTFOLIO_EXCEL = DATA_DIR / "excel" / "portfolio_tracker.xlsx"

# ── Color palette ─────────────────────────────────────────────
CLR_HEADER   = "1A3A5C"   # dark navy
CLR_BUY      = "C6EFCE"   # light green
CLR_SELL     = "FFCCCC"   # light red
CLR_HOLD     = "FFEB9C"   # light amber
CLR_PENDING  = "DAEEF3"   # light blue
CLR_CASH     = "F2F2F2"   # light grey
CLR_POSITIVE = "00B050"   # green font
CLR_NEGATIVE = "FF0000"   # red font
CLR_SUBHDR   = "BDD7EE"   # light blue subheader


def _hdr_style(ws, row: int, n_cols: int):
    """Apply navy header style to a row."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor=CLR_HEADER)
        cell.alignment = Alignment(horizontal='center',
                                   vertical='center', wrap_text=True)


def _pl_color(ws, row: int, col: int):
    """Color P&L cell green if positive, red if negative."""
    cell = ws.cell(row=row, column=col)
    try:
        val = float(cell.value or 0)
        if val > 0:
            cell.font = Font(color=CLR_POSITIVE, bold=True)
        elif val < 0:
            cell.font = Font(color=CLR_NEGATIVE, bold=True)
    except Exception:
        pass


def _action_color(ws, row: int, col: int):
    """Color the action cell based on action type."""
    cell  = ws.cell(row=row, column=col)
    val   = str(cell.value or '')
    color = None
    if 'BUY' in val or 'ACCUMULATE' in val:
        color = CLR_BUY
    elif 'SELL' in val:
        color = CLR_SELL
    elif 'PENDING' in val:
        color = CLR_PENDING
    elif 'SKIP' in val or 'CASH' in val:
        color = CLR_CASH
    if color:
        cell.fill = PatternFill("solid", fgColor=color)


def _auto_width(ws):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 35)


# ══════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ══════════════════════════════════════════════════════════════

def _build_daily_log(state: dict) -> pd.DataFrame:
    """Sheet 1: One row per position action per day."""
    rows = []
    for day in state.get('daily_log', []):
        for action in day.get('actions', []):
            # Parse action string: "BUY-FILLED SBIN qty=5 @ ₹820.00"
            parts  = action.split()
            act    = parts[0] if parts else ''
            sym    = parts[1] if len(parts) > 1 else ''

            # Find position data
            pos = state['positions'].get(sym, {})
            ct  = next((t for t in state.get('closed_trades', [])
                       if t['symbol'] == sym and
                       t['sell_date'] == day['date']), {})

            rows.append({
                'Date':           day['date'],
                'Day #':          day['day'],
                'Symbol':         sym,
                'Action':         act,
                'Entry Price ₹':  pos.get('entry_price') or ct.get('entry_price'),
                'Exit Price ₹':   ct.get('exit_price'),
                'Qty':            pos.get('qty') or ct.get('qty'),
                'Invested ₹':     pos.get('invested') or ct.get('invested'),
                'Current Value ₹': pos.get('current_value'),
                'P&L ₹':          ct.get('pl') or pos.get('unrealised_pl'),
                'P&L %':          ct.get('pl_pct') or pos.get('unrealised_pct'),
                'Stop Loss ₹':    pos.get('stop_loss'),
                'Stop Mode':      pos.get('stop_mode', 'TIGHT'),
                'Council Score':  pos.get('latest_score') or pos.get('entry_score'),
                'GPS':            pos.get('gps'),
                'Verdict':        pos.get('verdict') or ct.get('verdict', ''),
                'Hold Days':      pos.get('days_held') or ct.get('days_held'),
                'Sector':         pos.get('sector') or ct.get('sector', ''),
                'Reason':         ct.get('reason', '') or act,
                '1D Pred ₹':      pos.get('pred_1d'),
                '3D Pred ₹':      pos.get('pred_3d'),
                '7D Pred ₹':      pos.get('pred_7d'),
                'Day Mode':       day.get('day_mode', ''),
            })

    return pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=['Date', 'Day #', 'Symbol', 'Action', 'Entry Price ₹',
                 'Exit Price ₹', 'Qty', 'Invested ₹', 'Current Value ₹',
                 'P&L ₹', 'P&L %', 'Stop Loss ₹', 'Stop Mode',
                 'Council Score', 'GPS', 'Verdict', 'Hold Days',
                 'Sector', 'Reason', '1D Pred ₹', '3D Pred ₹',
                 '7D Pred ₹', 'Day Mode'])


def _build_daily_summary(state: dict) -> pd.DataFrame:
    """Sheet 2: One row per day — portfolio P&L summary."""
    rows = []
    prev_val = STARTING_CAPITAL

    for day in state.get('daily_log', []):
        pv    = day.get('portfolio_value_end', day.get('portfolio_value', prev_val))
        dp    = pv - prev_val
        dp_pct = dp / prev_val * 100 if prev_val else 0

        # Best/worst stock today
        # FIX: With only 1 closed trade, best and worst become the same
        # stock (misleading). Only populate when 2+ trades exist that day.
        today_trades = [t for t in state.get('closed_trades', [])
                       if t.get('sell_date') == day['date']]

        if len(today_trades) >= 2:
            best  = max(today_trades, key=lambda t: t.get('pl_pct', 0)).get('symbol', '')
            worst = min(today_trades, key=lambda t: t.get('pl_pct', 0)).get('symbol', '')
        elif len(today_trades) == 1:
            # Only one trade — label it neutrally, not best/worst
            only_sym = today_trades[0].get('symbol', '')
            only_pl  = today_trades[0].get('pl_pct', 0)
            if only_pl >= 0:
                best, worst = only_sym, ''
            else:
                best, worst = '', only_sym
        else:
            best, worst = '', ''

        rows.append({
            'Date':              day['date'],
            'Day #':             day['day'],
            'Portfolio Value ₹': round(pv, 2),
            'Cash ₹':            round(day.get('cash', 0), 2),
            'Invested ₹':        round(pv - day.get('cash', 0), 2),
            'Daily P&L ₹':       round(dp, 2),
            'Daily P&L %':       round(dp_pct, 2),
            'Total P&L ₹':       round(pv - STARTING_CAPITAL, 2),
            'Total Return %':    round((pv - STARTING_CAPITAL)
                                       / STARTING_CAPITAL * 100, 2),
            'Deployed Tonight ₹': sum(
                float(a.split('₹')[-1].replace(',', ''))
                for a in day.get('actions', [])
                if 'PENDING' in a and '₹' in a
            ),
            'Stocks Held':       day.get('positions_held', 0),
            'Buys Today':        day.get('buys', 0),
            'Sells Today':       day.get('sells', 0),
            'Avg GPS':           day.get('avg_gps', 0),
            'Market Score':      day.get('market_score', 0),
            'Day Mode':          day.get('day_mode', ''),
            'Best Stock':        best,
            'Worst Stock':       worst,
            'Skipped':           day.get('skipped', False),
        })
        prev_val = pv

    return pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=['Date', 'Day #', 'Portfolio Value ₹', 'Cash ₹',
                 'Invested ₹', 'Daily P&L ₹', 'Daily P&L %',
                 'Total P&L ₹', 'Total Return %', 'Deployed Tonight ₹',
                 'Stocks Held', 'Buys Today', 'Sells Today',
                 'Avg GPS', 'Market Score', 'Day Mode',
                 'Best Stock', 'Worst Stock', 'Skipped'])


def _build_open_positions(state: dict) -> pd.DataFrame:
    """Sheet 3: Live snapshot of all open positions."""
    rows = []
    for sym, pos in state.get('positions', {}).items():
        score_hist = pos.get('score_history', [])
        if len(score_hist) >= 2:
            trend = ('↑' if score_hist[-1] > score_hist[-2] else
                     '↓' if score_hist[-1] < score_hist[-2] else '→')
        else:
            trend = '—'

        rows.append({
            'Symbol':           sym,
            'Entry Date':       pos.get('entry_date'),
            'Entry Price ₹':    pos.get('entry_price'),
            'Current Price ₹':  pos.get('current_price'),
            'Qty':              pos.get('qty'),
            'Invested ₹':       round(pos.get('invested', 0), 2),
            'Current Value ₹':  round(pos.get('current_value', 0), 2),
            'Unrealised P&L ₹': round(pos.get('unrealised_pl', 0), 2),
            'Unrealised P&L %': round(pos.get('unrealised_pct', 0), 2),
            'Stop Loss ₹':      pos.get('stop_loss'),
            'Stop Mode':        pos.get('stop_mode', 'TIGHT'),
            'Days Held':        pos.get('days_held', 0),
            'Entry Score':      pos.get('entry_score'),
            'Latest Score':     pos.get('latest_score'),
            'Score Trend':      trend,
            'Improving Nights': pos.get('improving_nights', 0),
            'Sector':           pos.get('sector', ''),
            'Verdict':          pos.get('verdict', ''),
            'GPS':              pos.get('gps'),
            '21D Target ₹':     pos.get('pred_21d'),
            'Google Finance':   f"https://www.google.com/finance/quote/{sym}:NSE",
        })

    return pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=['Symbol', 'Entry Date', 'Entry Price ₹',
                 'Current Price ₹', 'Qty', 'Invested ₹',
                 'Current Value ₹', 'Unrealised P&L ₹',
                 'Unrealised P&L %', 'Stop Loss ₹', 'Stop Mode',
                 'Days Held', 'Entry Score', 'Latest Score',
                 'Score Trend', 'Sector', 'Verdict'])


def _build_closed_trades(state: dict) -> pd.DataFrame:
    """Sheet 4: All completed trades."""
    rows = []
    for t in state.get('closed_trades', []):
        rows.append({
            'Symbol':        t.get('symbol'),
            'Buy Date':      t.get('buy_date'),
            'Sell Date':     t.get('sell_date'),
            'Entry ₹':       t.get('entry_price'),
            'Exit ₹':        t.get('exit_price'),
            'Qty':           t.get('qty'),
            'Invested ₹':    round(t.get('invested', 0), 2),
            'Proceeds ₹':    round(t.get('proceeds', 0), 2),
            'Realised P&L ₹': round(t.get('pl', 0), 2),
            'Return %':      round(t.get('pl_pct', 0), 2),
            'Sell Reason':   t.get('reason'),
            'Days Held':     t.get('days_held'),
            'Entry Score':   t.get('entry_score'),
            'Exit Score':    t.get('exit_score'),
            'Sector':        t.get('sector'),
            'Win/Loss':      '✅ WIN' if t.get('pl', 0) > 0 else '❌ LOSS',
        })

    return pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=['Symbol', 'Buy Date', 'Sell Date', 'Entry ₹', 'Exit ₹',
                 'Qty', 'Invested ₹', 'Proceeds ₹', 'Realised P&L ₹',
                 'Return %', 'Sell Reason', 'Days Held',
                 'Entry Score', 'Exit Score', 'Sector', 'Win/Loss'])


def _build_learning_feedback(state: dict) -> pd.DataFrame:
    """Sheet 5: Weight change audit trail."""
    rows = []
    for change in state.get('weight_changes', []):
        rows.append({
            'Date':           change.get('date'),
            'Parameter':      change.get('param'),
            'Old Value':      change.get('old'),
            'New Value':      change.get('new'),
            'Change':         round(float(change.get('new', 0)) -
                                    float(change.get('old', 0)), 4),
            'Reason':         change.get('reason'),
            'Triggered By':   change.get('trigger', 'weekly_review'),
        })

    return pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=['Date', 'Parameter', 'Old Value', 'New Value',
                 'Change', 'Reason', 'Triggered By'])


# ══════════════════════════════════════════════════════════════
# MAIN WRITE FUNCTION
# ══════════════════════════════════════════════════════════════

def save_portfolio_excel(state: dict = None,
                          print_output: bool = True,
                          max_retries: int = 3,
                          retry_wait_s: int = 15) -> Path:
    """
    Build and save portfolio_tracker.xlsx.
    Called every night from night_runner.py after portfolio engine runs.

    FIX: Retries on PermissionError (file open in Excel, or
    transient Windows file lock) up to max_retries times with
    a wait between attempts. If all retries fail, falls back to
    a timestamped backup file so tonight's data is never lost --
    it just needs manual merging later instead of being silently
    dropped (the bug that caused Day 2 and Day 5 to go missing
    from the Excel despite the underlying data being correct).
    """
    if state is None:
        state = load_state()

    last_error = None
    for attempt in range(1, max_retries + 1):
        try:
            return _write_portfolio_excel(state, print_output, PORTFOLIO_EXCEL)
        except PermissionError as e:
            last_error = e
            if print_output:
                print(f"  [PORTFOLIO EXCEL] File locked (attempt "
                      f"{attempt}/{max_retries}) -- is it open in Excel?")
            if attempt < max_retries:
                if print_output:
                    print(f"  Retrying in {retry_wait_s}s... "
                          f"(close the file in Excel if you have it open)")
                time.sleep(retry_wait_s)
        except Exception as e:
            last_error = e
            break

    # All retries exhausted (or non-permission error) -- save to
    # a timestamped backup so tonight's data is never silently lost
    if print_output:
        print(f"  [PORTFOLIO EXCEL] Could not write to main file after "
              f"{max_retries} attempts: {last_error}")
        print(f"  Saving to timestamped backup instead...")

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_path = PORTFOLIO_EXCEL.parent / f"portfolio_tracker_BACKUP_{ts}.xlsx"
    try:
        result = _write_portfolio_excel(state, print_output, backup_path)
        if print_output:
            print(f"  ⚠ Saved to backup: {backup_path.name}")
            print(f"  ⚠ ACTION NEEDED: close portfolio_tracker.xlsx in Excel, "
                  f"then run this to merge tonight's data:")
            print(f"     python -c \"from portfolio_engine import load_state; "
                  f"from portfolio_excel import save_portfolio_excel; "
                  f"save_portfolio_excel(load_state(), print_output=True)\"")
        return backup_path
    except Exception as e2:
        print(f"  [PORTFOLIO EXCEL] Backup save also failed: {e2}")
        import traceback
        traceback.print_exc()
        return None


def _write_portfolio_excel(state: dict, print_output: bool,
                           output_path: Path) -> Path:
    """
    Actual Excel write logic, separated out so save_portfolio_excel()
    can retry it against either the main file or a backup path.
    """
    try:
        df_log      = _build_daily_log(state)
        df_summary  = _build_daily_summary(state)
        df_open     = _build_open_positions(state)
        df_closed   = _build_closed_trades(state)
        df_learning = _build_learning_feedback(state)

        with pd.ExcelWriter(str(output_path),
                            engine='openpyxl') as writer:

            # ── Sheet 1: Daily Log ─────────────────────────────
            df_log.to_excel(writer, sheet_name='Daily Log', index=False)
            ws = writer.sheets['Daily Log']
            _hdr_style(ws, 1, len(df_log.columns))
            ws.freeze_panes = 'A2'
            for row_idx in range(2, len(df_log) + 2):
                action_col = df_log.columns.get_loc('Action') + 1
                _action_color(ws, row_idx, action_col)
                pl_col = df_log.columns.get_loc('P&L ₹') + 1
                _pl_color(ws, row_idx, pl_col)
            _auto_width(ws)

            # ── Sheet 2: Daily Summary ─────────────────────────
            df_summary.to_excel(writer, sheet_name='Daily Summary', index=False)
            ws2 = writer.sheets['Daily Summary']
            _hdr_style(ws2, 1, len(df_summary.columns))
            ws2.freeze_panes = 'A2'
            for row_idx in range(2, len(df_summary) + 2):
                for col_name in ['Daily P&L ₹', 'Daily P&L %',
                                 'Total P&L ₹', 'Total Return %']:
                    if col_name in df_summary.columns:
                        _pl_color(ws2, row_idx,
                                  df_summary.columns.get_loc(col_name) + 1)
            _auto_width(ws2)

            # ── Sheet 3: Open Positions ────────────────────────
            df_open.to_excel(writer, sheet_name='Open Positions', index=False)
            ws3 = writer.sheets['Open Positions']
            _hdr_style(ws3, 1, len(df_open.columns))
            ws3.freeze_panes = 'A2'
            for row_idx in range(2, len(df_open) + 2):
                for col_name in ['Unrealised P&L ₹', 'Unrealised P&L %']:
                    if col_name in df_open.columns:
                        _pl_color(ws3, row_idx,
                                  df_open.columns.get_loc(col_name) + 1)
            _auto_width(ws3)

            # ── Sheet 4: Closed Trades ─────────────────────────
            df_closed.to_excel(writer, sheet_name='Closed Trades', index=False)
            ws4 = writer.sheets['Closed Trades']
            _hdr_style(ws4, 1, len(df_closed.columns))
            ws4.freeze_panes = 'A2'
            for row_idx in range(2, len(df_closed) + 2):
                for col_name in ['Realised P&L ₹', 'Return %']:
                    if col_name in df_closed.columns:
                        _pl_color(ws4, row_idx,
                                  df_closed.columns.get_loc(col_name) + 1)
            _auto_width(ws4)

            # ── Sheet 5: Learning Feedback ─────────────────────
            df_learning.to_excel(writer, sheet_name='Learning Feedback',
                                 index=False)
            ws5 = writer.sheets['Learning Feedback']
            _hdr_style(ws5, 1, len(df_learning.columns))
            _auto_width(ws5)

        if print_output:
            pv  = _portfolio_value(state)
            pl  = pv - STARTING_CAPITAL
            day = state.get('day', 0)
            print(f"\n  📊 Portfolio Excel saved: {output_path.name}")
            print(f"     Day {day}/21 | ₹{pv:,.0f} | "
                  f"P&L: ₹{pl:+,.0f} ({pl/STARTING_CAPITAL*100:+.1f}%)")
            print(f"     Open: {len(state['positions'])} | "
                  f"Closed: {len(state.get('closed_trades',[]))} trades")

        return output_path

    except PermissionError:
        raise   # let save_portfolio_excel() handle retry logic
    except Exception as e:
        print(f"  [PORTFOLIO EXCEL] Error: {e}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    state = load_state()
    save_portfolio_excel(state)
    print(f"  Saved to: {PORTFOLIO_EXCEL}")