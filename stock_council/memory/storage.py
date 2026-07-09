# ============================================================
# memory/storage.py
# Local file storage for ALL market data
# ============================================================
# SAVES TO:
#   data/prices/{SYMBOL}.csv          — full OHLCV history
#   data/fundamentals/{SYMBOL}.json   — fundamentals snapshot
#   data/news/{SYMBOL}/{DATE}.txt     — news text per day
#   data/excel/master_{DATE}.xlsx     — master Excel with all stocks
#   data/excel/sectors_{DATE}.xlsx    — sector rankings Excel
#   data/scores/{SYMBOL}.json         — historical bot scores
#   memory/market_{DATE}.json         — full daily market snapshot
#   memory/council_{DATE}.json        — bot council session results
#
# LOADING:
#   On startup, load all local files first
#   Only hit internet for data NOT in local store or stale
# ============================================================

import os
import json
import csv
import shutil
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime, timedelta
import pytz

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from config import BASE_DIR, DATA_DIR, REPORTS_DIR, VERBOSE_DEBUG

IST = pytz.timezone('Asia/Kolkata')

# ── Paths ─────────────────────────────────────────────────────
PRICES_DIR      = DATA_DIR / "prices"
FUND_DIR        = DATA_DIR / "fundamentals"
NEWS_DIR        = DATA_DIR / "news"
EXCEL_DIR       = DATA_DIR / "excel"
SCORES_DIR      = DATA_DIR / "scores"
MEMORY_DIR      = BASE_DIR / "memory"
VECTORS_DIR     = DATA_DIR / "vectors"

for d in [PRICES_DIR, FUND_DIR, NEWS_DIR, EXCEL_DIR,
          SCORES_DIR, MEMORY_DIR, VECTORS_DIR]:
    d.mkdir(parents=True, exist_ok=True)


def today_str() -> str:
    return datetime.now(IST).strftime('%Y-%m-%d')

def now_str() -> str:
    return datetime.now(IST).strftime('%Y-%m-%d_%H-%M')


# ══════════════════════════════════════════════════════════════
# PRICE DATA  →  data/prices/{SYMBOL}.csv
# ══════════════════════════════════════════════════════════════

def save_prices_csv(symbol: str, df: pd.DataFrame):
    """
    Save OHLCV dataframe to CSV.
    Appends new rows, no duplicates.
    CSV columns: Date, Open, High, Low, Close, Volume
    """
    path = PRICES_DIR / f"{symbol.replace('.NS','').replace('.BO','')}.csv"
    df_save = df.copy()
    df_save.index = pd.to_datetime(df_save.index)
    # Strip timezone info to keep everything tz-naive
    if df_save.index.tz is not None:
        df_save.index = df_save.index.tz_localize(None)
    df_save.index.name = 'Date'
    # Keep only OHLCV columns that exist
    cols = [c for c in ['Open','High','Low','Close','Volume'] if c in df_save.columns]
    df_save = df_save[cols]

    if path.exists():
        existing = pd.read_csv(path, index_col='Date', parse_dates=True)
        # Strip timezone from existing too
        if existing.index.tz is not None:
            existing.index = existing.index.tz_localize(None)
        combined = pd.concat([existing, df_save])
        combined = combined[~combined.index.duplicated(keep='last')]
        combined.sort_index(inplace=True)
        combined.to_csv(path)
    else:
        df_save.to_csv(path)

    if VERBOSE_DEBUG:
        print(f"[STORE] Saved prices: {path.name}")


def load_prices_csv(symbol: str, days: int = None) -> pd.DataFrame | None:
    """
    Load OHLCV from local CSV.
    Returns None if file doesn't exist.
    """
    path = PRICES_DIR / f"{symbol.replace('.NS','').replace('.BO','')}.csv"
    if not path.exists():
        return None
    df = pd.read_csv(path, index_col='Date', parse_dates=True)
    df.sort_index(inplace=True)
    if days:
        # FIX: datetime.now() includes a time-of-day component
        # (e.g. 2026-06-22 14:30:00). A date-only CSV row like
        # "2026-06-19 00:00:00" can fall BEFORE a cutoff of
        # "2026-06-19 14:30:00", silently excluding the most
        # recent trading day from the result. Normalize cutoff
        # to midnight so date-only comparisons work correctly.
        cutoff = (datetime.now() - timedelta(days=days)).replace(
            hour=0, minute=0, second=0, microsecond=0)
        if df.index.tz is not None:
            df.index = df.index.tz_localize(None)
        df = df[df.index >= cutoff]
    return df if not df.empty else None


def is_prices_fresh(symbol: str, max_age_hours: float = 6.0) -> bool:
    """Check if local price CSV is recent enough."""
    path = PRICES_DIR / f"{symbol.replace('.NS','').replace('.BO','')}.csv"
    if not path.exists():
        return False
    age_hours = (datetime.now().timestamp() - path.stat().st_mtime) / 3600
    return age_hours < max_age_hours


# ══════════════════════════════════════════════════════════════
# FUNDAMENTALS  →  data/fundamentals/{SYMBOL}.json
# ══════════════════════════════════════════════════════════════

def save_fundamentals_json(symbol: str, data: dict):
    """Save fundamentals dict as JSON with timestamp."""
    sym = symbol.replace('.NS','').replace('.BO','')
    path = FUND_DIR / f"{sym}.json"
    data['_saved_at'] = datetime.now(IST).isoformat()
    data['_date'] = today_str()
    with open(path, 'w') as f:
        json.dump(data, f, indent=2, default=str)
    if VERBOSE_DEBUG:
        print(f"[STORE] Saved fundamentals: {path.name}")


def load_fundamentals_json(symbol: str, max_age_hours: float = 24.0) -> dict | None:
    """Load fundamentals from local JSON if fresh enough."""
    sym = symbol.replace('.NS','').replace('.BO','')
    path = FUND_DIR / f"{sym}.json"
    if not path.exists():
        return None
    age_hours = (datetime.now().timestamp() - path.stat().st_mtime) / 3600
    if age_hours > max_age_hours:
        return None
    with open(path) as f:
        return json.load(f)


# ══════════════════════════════════════════════════════════════
# NEWS  →  data/news/{SYMBOL}/{DATE}.txt
# ══════════════════════════════════════════════════════════════

def save_news_text(symbol: str, articles: list, date: str = None):
    """
    Save news articles as a plain text file per stock per day.
    Format: one article per block, human-readable.
    Also saves machine-readable JSON alongside.
    """
    sym = symbol.replace('.NS','').replace('.BO','')
    date = date or today_str()
    stock_news_dir = NEWS_DIR / sym
    stock_news_dir.mkdir(exist_ok=True)

    # Human-readable text
    txt_path = stock_news_dir / f"{date}.txt"
    lines = [
        f"NEWS FOR: {sym}",
        f"Date:     {date}",
        f"Articles: {len(articles)}",
        "=" * 60,
        ""
    ]
    for i, a in enumerate(articles, 1):
        sent = a.get('sentiment', 0)
        sent_label = "POSITIVE" if sent > 0.05 else "NEGATIVE" if sent < -0.05 else "NEUTRAL"
        lines += [
            f"[{i}] {a.get('title','')}",
            f"    Source:    {a.get('source','')}",
            f"    Published: {a.get('published','')}",
            f"    Sentiment: {sent:.4f} ({sent_label})",
            f"    Summary:   {a.get('summary','')[:300]}",
            f"    URL:       {a.get('url','')}",
            "",
        ]
    txt_path.write_text('\n'.join(lines), encoding='utf-8')

    # JSON for machine reading
    json_path = stock_news_dir / f"{date}.json"
    with open(json_path, 'w') as f:
        json.dump({'date': date, 'symbol': sym, 'articles': articles}, f, indent=2, default=str)

    if VERBOSE_DEBUG:
        print(f"[STORE] Saved news: {txt_path.name}")


def load_news_text(symbol: str, date: str = None, days_back: int = 1) -> list:
    """
    Load saved news articles from local files.
    Returns list of article dicts.
    """
    sym = symbol.replace('.NS','').replace('.BO','')
    stock_news_dir = NEWS_DIR / sym
    if not stock_news_dir.exists():
        return []

    articles = []
    for d in range(days_back):
        check_date = (datetime.now() - timedelta(days=d)).strftime('%Y-%m-%d')
        json_path = stock_news_dir / f"{check_date}.json"
        if json_path.exists():
            with open(json_path) as f:
                data = json.load(f)
                articles.extend(data.get('articles', []))
    return articles


def load_news_as_context(symbol: str, days_back: int = 7) -> str:
    """
    Load last N days of news as a single text block.
    Used to feed context to LLM.
    """
    sym = symbol.replace('.NS','').replace('.BO','')
    stock_news_dir = NEWS_DIR / sym
    if not stock_news_dir.exists():
        return f"No saved news for {sym}"

    blocks = []
    for d in range(days_back):
        check_date = (datetime.now() - timedelta(days=d)).strftime('%Y-%m-%d')
        txt_path = stock_news_dir / f"{check_date}.txt"
        if txt_path.exists():
            blocks.append(txt_path.read_text(encoding='utf-8'))

    return '\n\n'.join(blocks) if blocks else f"No local news for {sym}"


# ══════════════════════════════════════════════════════════════
# BOT SCORES HISTORY  →  data/scores/{SYMBOL}.json
# ══════════════════════════════════════════════════════════════

def save_bot_scores(symbol: str, date: str, scores: dict, verdict: str,
                    composite: float, analysis_texts: dict = None):
    """
    Save complete bot council session results for a stock.
    Appends to history — so we accumulate all past sessions.

    scores = {
        'fundamental': 7.2, 'technical': 6.8,
        'news': 7.0, 'sentiment': 6.5, 'risk': 4.0
    }
    analysis_texts = {
        'fundamental': "...", 'technical': "...", ...
    }
    """
    sym = symbol.replace('.NS','').replace('.BO','')
    path = SCORES_DIR / f"{sym}.json"

    # Load existing history
    history = []
    if path.exists():
        with open(path) as f:
            history = json.load(f)

    entry = {
        'date': date,
        'timestamp': datetime.now(IST).isoformat(),
        'scores': scores,
        'verdict': verdict,
        'composite': composite,
    }
    if analysis_texts:
        entry['analysis'] = {k: v[:500] for k, v in analysis_texts.items()}

    history.append(entry)

    # Keep last 90 days
    history = history[-90:]

    with open(path, 'w') as f:
        json.dump(history, f, indent=2, default=str)


def load_bot_scores_history(symbol: str, days: int = 30) -> list:
    """
    Load past bot scores for a stock.
    Used by bots to see their own trend.
    """
    sym = symbol.replace('.NS','').replace('.BO','')
    path = SCORES_DIR / f"{sym}.json"
    if not path.exists():
        return []
    with open(path) as f:
        history = json.load(f)
    cutoff = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
    return [h for h in history if h.get('date', '') >= cutoff]


def get_score_trend(symbol: str, bot_name: str, days: int = 14) -> dict:
    """
    Get score trend for one bot over N days.
    Returns: {trend: 'improving'/'declining'/'stable', avg: 6.2, delta: +0.8}
    """
    history = load_bot_scores_history(symbol, days=days)
    if len(history) < 2:
        return {'trend': 'insufficient_data', 'avg': None, 'delta': None}

    scores = [h['scores'].get(bot_name) for h in history
              if h.get('scores', {}).get(bot_name) is not None]
    if len(scores) < 2:
        return {'trend': 'insufficient_data', 'avg': None, 'delta': None}

    avg = round(sum(scores) / len(scores), 2)
    delta = round(scores[-1] - scores[0], 2)
    trend = 'improving' if delta > 0.5 else 'declining' if delta < -0.5 else 'stable'
    return {'trend': trend, 'avg': avg, 'delta': delta, 'values': scores}


# ══════════════════════════════════════════════════════════════
# MARKET MEMORY  →  memory/market_{DATE}.json
# ══════════════════════════════════════════════════════════════

def save_market_snapshot(snapshot_dict: dict, date: str = None):
    """Save full market snapshot (Level 1) to memory folder."""
    date = date or today_str()
    path = MEMORY_DIR / f"market_{date}.json"
    snapshot_dict['_date'] = date
    snapshot_dict['_saved_at'] = datetime.now(IST).isoformat()
    with open(path, 'w') as f:
        json.dump(snapshot_dict, f, indent=2, default=str)
    if VERBOSE_DEBUG:
        print(f"[STORE] Saved market snapshot: {path.name}")


def load_market_snapshot(date: str = None) -> dict | None:
    """Load saved market snapshot."""
    date = date or today_str()
    path = MEMORY_DIR / f"market_{date}.json"
    if path.exists():
        with open(path) as f:
            return json.load(f)
    return None


def save_council_session(date: str, stocks_results: list, sector_results: list):
    """Save full council session — all stocks, all sectors, all verdicts."""
    path = MEMORY_DIR / f"council_{date}_{now_str().split('_')[1]}.json"
    session = {
        'date': date,
        'timestamp': datetime.now(IST).isoformat(),
        'sectors': [
            {
                'name': s.name,
                'score': s.score,
                'verdict': s.verdict,
                'avg_1m': getattr(s, 'avg_1m_change', None),
                'avg_rsi': getattr(s, 'avg_rsi', None),
                'top_stocks': [t[0] for t in getattr(s, 'top_stocks', [])],
            }
            for s in sector_results
        ],
        'stocks': [
            {
                'symbol': v.symbol,
                'sector': v.sector,
                'price': v.current_price,
                'verdict': v.verdict,
                'score': v.final_score,
                'scores': {
                    'fundamental': v.fundamental_score,
                    'technical': v.technical_score,
                    'news': v.news_score,
                    'sentiment': v.sentiment_score,
                    'risk': v.risk_score,
                },
                'change_1m': v.change_1m,
                'rsi': v.rsi,
            }
            for v in stocks_results
        ]
    }
    with open(path, 'w') as f:
        json.dump(session, f, indent=2, default=str)
    return path


def load_last_council_session() -> dict | None:
    """Load the most recent council session."""
    sessions = sorted(MEMORY_DIR.glob('council_*.json'), reverse=True)
    if not sessions:
        return None
    with open(sessions[0]) as f:
        return json.load(f)


def load_yesterdays_verdicts() -> dict:
    """
    Load yesterday's stock verdicts.
    Used by bots at market open to see what changed overnight.
    Returns: {symbol: {verdict, score, scores, price}}
    """
    yesterday = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    sessions = sorted(MEMORY_DIR.glob(f'council_{yesterday}*.json'), reverse=True)

    # If no yesterday, try last available
    if not sessions:
        sessions = sorted(MEMORY_DIR.glob('council_*.json'), reverse=True)

    if not sessions:
        return {}

    with open(sessions[0]) as f:
        data = json.load(f)

    return {s['symbol']: s for s in data.get('stocks', [])}


# ══════════════════════════════════════════════════════════════
# EXCEL REPORTS  →  data/excel/
# ══════════════════════════════════════════════════════════════

def save_master_excel(stocks_results: list, sector_results: list,
                       market_snap_dict: dict, date: str = None):
    """
    Save complete daily Excel report with multiple sheets:
      Sheet 1: Master Stock Rankings
      Sheet 2: Sector Rankings
      Sheet 3: Market Overview
      Sheet 4: Hot Sectors × Hot Stocks matrix
      Sheet 5: Score History (last 30 days composite trend)
    """
    date = date or today_str()
    path = EXCEL_DIR / f"market_{date}.xlsx"

    try:
        with pd.ExcelWriter(str(path), engine='openpyxl') as writer:

            # ── Sheet 1: All Stock Rankings ────────────────────
            stock_rows = []
            for v in stocks_results:
                history = load_bot_scores_history(v.symbol, days=7)
                prev_score = history[-2]['composite'] if len(history) >= 2 else None
                score_change = round(v.final_score - prev_score, 2) if prev_score else None

                stock_rows.append({
                    'Rank':          v.rank,
                    'Symbol':        v.symbol,
                    'Company':       getattr(v, 'company_name', v.symbol),
                    'Sector':        v.sector,
                    'Verdict':       v.verdict,
                    'Score':         v.final_score,
                    'Score Change':  score_change,
                    'Price (₹)':     v.current_price,
                    '1D Change %':   v.change_1d,
                    '1M Change %':   v.change_1m,
                    'RSI':           v.rsi,
                    'Volume Ratio':  v.volume_ratio,
                    'Above 50DMA':   v.above_50dma,
                    'P/E Ratio':     v.pe_ratio,
                    'ROE %':         v.roe,
                    'Beta':          v.beta,
                    'Fund Score':    v.fundamental_score,
                    'Tech Score':    v.technical_score,
                    'News Score':    v.news_score,
                    'Sent Score':    v.sentiment_score,
                    'Risk Level':    v.risk_score,
                    'LLM Summary':   (v.llm_analysis or '')[:200],
                })

            df_stocks = pd.DataFrame(stock_rows)
            df_stocks.to_excel(writer, sheet_name='Stock Rankings', index=False)

            # Apply conditional formatting
            ws = writer.sheets['Stock Rankings']
            _format_excel_sheet(ws, df_stocks)

            # ── Sheet 2: Sector Rankings ───────────────────────
            sector_rows = []
            for s in sector_results:
                sector_rows.append({
                    'Rank':            s.rank,
                    'Sector':          s.name,
                    'Verdict':         s.verdict,
                    'Score':           s.score,
                    'Avg 1D Change %': getattr(s, 'avg_1d_change', None),
                    'Avg 1M Change %': getattr(s, 'avg_1m_change', None),
                    'vs Nifty 1M %':   getattr(s, 'vs_nifty_1m', None),
                    'Avg RSI':         getattr(s, 'avg_rsi', None),
                    'Avg Vol Ratio':   getattr(s, 'avg_volume_ratio', None),
                    '% Above 50DMA':   getattr(s, 'pct_above_50dma', None),
                    '% Above 200DMA':  getattr(s, 'pct_above_200dma', None),
                    'Avg PE':          getattr(s, 'avg_pe', None),
                    'Avg ROE':         getattr(s, 'avg_roe', None),
                    'Momentum Score':  getattr(s, 'momentum_score', None),
                    'Breadth Score':   getattr(s, 'breadth_score', None),
                    'Technical Score': getattr(s, 'technical_score', None),
                    'Macro Score':     getattr(s, 'macro_score', None),
                    'Top Stocks':      ', '.join(t[0] for t in getattr(s, 'top_stocks', [])),
                    'Analysis':        (s.llm_analysis or '')[:300],
                })
            df_sectors = pd.DataFrame(sector_rows)
            df_sectors.to_excel(writer, sheet_name='Sector Rankings', index=False)

            # ── Sheet 3: Market Overview ───────────────────────
            mkt_rows = [
                {'Metric': 'Date', 'Value': date},
                {'Metric': 'Market Score', 'Value': market_snap_dict.get('market_score')},
                {'Metric': 'Market Outlook', 'Value': market_snap_dict.get('market_outlook')},
                {'Metric': 'FII Net (₹ Cr)', 'Value': market_snap_dict.get('fii_net_cr')},
                {'Metric': 'DII Net (₹ Cr)', 'Value': market_snap_dict.get('dii_net_cr')},
                {'Metric': 'FII/DII Signal', 'Value': market_snap_dict.get('fii_dii_signal')},
                {'Metric': 'India VIX', 'Value': market_snap_dict.get('vix')},
                {'Metric': 'VIX Signal', 'Value': market_snap_dict.get('vix_signal')},
                {'Metric': 'Advances', 'Value': market_snap_dict.get('advance_count')},
                {'Metric': 'Declines', 'Value': market_snap_dict.get('decline_count')},
                {'Metric': 'A/D Ratio', 'Value': market_snap_dict.get('ad_ratio')},
                {'Metric': 'Breadth Signal', 'Value': market_snap_dict.get('breadth_signal')},
            ]
            # Add indices
            for idx in market_snap_dict.get('indices', []):
                mkt_rows.append({
                    'Metric': f"Index: {idx.get('name','')}",
                    'Value': f"{idx.get('last','')} ({idx.get('pct_change',0):+.2f}%)"
                })
            # Add global cues
            for name, data in market_snap_dict.get('global_cues', {}).items():
                mkt_rows.append({
                    'Metric': f"Global: {name}",
                    'Value': f"{data.get('price','')} ({data.get('change_pct',0):+.2f}%)"
                })
            pd.DataFrame(mkt_rows).to_excel(writer, sheet_name='Market Overview', index=False)

            # ── Sheet 4: Hot Sectors × Stocks Matrix ──────────
            hot_sectors = [s for s in sector_results if s.rank <= 5]
            matrix_rows = []
            for s in hot_sectors:
                row = {'Sector': s.name, 'Sector Score': s.score, 'Verdict': s.verdict}
                # Add top 5 stocks from this sector
                stocks_in_sector = [v for v in stocks_results if v.sector == s.name]
                for i, sv in enumerate(stocks_in_sector[:6], 1):
                    row[f'Stock {i}'] = sv.symbol
                    row[f'Stock {i} Score'] = sv.final_score
                    row[f'Stock {i} Verdict'] = sv.verdict
                matrix_rows.append(row)
            if matrix_rows:
                pd.DataFrame(matrix_rows).to_excel(
                    writer, sheet_name='Hot Sectors Matrix', index=False)

            # ── Sheet 5: Score History ─────────────────────────
            top_symbols = [v.symbol for v in stocks_results[:20]]
            history_rows = []
            for sym in top_symbols:
                hist = load_bot_scores_history(sym, days=30)
                for h in hist:
                    history_rows.append({
                        'Symbol':    sym,
                        'Date':      h.get('date'),
                        'Verdict':   h.get('verdict'),
                        'Composite': h.get('composite'),
                        'Fundamental': h.get('scores', {}).get('fundamental'),
                        'Technical':   h.get('scores', {}).get('technical'),
                        'News':        h.get('scores', {}).get('news'),
                        'Sentiment':   h.get('scores', {}).get('sentiment'),
                        'Risk':        h.get('scores', {}).get('risk'),
                    })
            if history_rows:
                pd.DataFrame(history_rows).to_excel(
                    writer, sheet_name='Score History', index=False)

        if VERBOSE_DEBUG:
            print(f"[STORE] Saved Excel: {path.name}")
        return path

    except Exception as e:
        print(f"[STORE] Excel save error: {e}")
        return None


def _format_excel_sheet(ws, df: pd.DataFrame):
    """Apply basic Excel formatting — column widths, bold headers."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        # Bold headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Column widths
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max() or 0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 30)

        # Color verdict column
        verdict_col = None
        for i, col in enumerate(df.columns, 1):
            if col == 'Verdict':
                verdict_col = i
                break

        if verdict_col:
            COLORS = {
                'STRONG BUY': 'FF008000',
                'BUY':        'FF90EE90',
                'ACCUMULATE': 'FFADFF2F',
                'HOLD':       'FFFFFF00',
                'REDUCE':     'FFFFA500',
                'SELL':       'FFFF6347',
                'STRONG SELL':'FFFF0000',
            }
            for row in ws.iter_rows(min_row=2, min_col=verdict_col, max_col=verdict_col):
                for cell in row:
                    val = str(cell.value or '').upper().split()[0] + \
                          (' ' + str(cell.value or '').upper().split()[1]
                           if len(str(cell.value or '').split()) > 1 else '')
                    for verdict_key, color in COLORS.items():
                        if verdict_key in (cell.value or '').upper():
                            cell.fill = PatternFill(fill_type='solid', fgColor=color)
                            break
    except Exception:
        pass  # formatting is optional


# ══════════════════════════════════════════════════════════════
# UTILITY FUNCTIONS
# ══════════════════════════════════════════════════════════════

def list_stored_symbols() -> list:
    """List all symbols that have local price CSVs."""
    return [f.stem for f in sorted(PRICES_DIR.glob('*.csv'))]


def get_storage_summary() -> dict:
    """Summary of what's stored locally."""
    price_files = list(PRICES_DIR.glob('*.csv'))
    fund_files = list(FUND_DIR.glob('*.json'))
    news_folders = list(NEWS_DIR.iterdir()) if NEWS_DIR.exists() else []
    excel_files = list(EXCEL_DIR.glob('*.xlsx'))
    score_files = list(SCORES_DIR.glob('*.json'))
    council_files = list(MEMORY_DIR.glob('council_*.json'))

    # Total size
    total_bytes = sum(f.stat().st_size for f in
                      list(PRICES_DIR.glob('**/*')) +
                      list(FUND_DIR.glob('**/*')) +
                      list(NEWS_DIR.glob('**/*')) +
                      list(EXCEL_DIR.glob('**/*')) +
                      list(SCORES_DIR.glob('**/*')) +
                      list(MEMORY_DIR.glob('**/*'))
                      if f.is_file())

    return {
        'price_csvs':       len(price_files),
        'fundamental_jsons': len(fund_files),
        'news_stock_dirs':   len(news_folders),
        'excel_reports':     len(excel_files),
        'score_histories':   len(score_files),
        'council_sessions':  len(council_files),
        'total_size_mb':     round(total_bytes / 1024 / 1024, 2),
        'oldest_excel':      min((f.name for f in excel_files), default='none'),
        'latest_excel':      max((f.name for f in excel_files), default='none'),
    }


def cleanup_old_files(keep_days: int = 90):
    """
    Delete files older than keep_days to save disk space.
    Keeps: price CSVs (always), fundamentals (always)
    Cleans: old news text, old council JSON, old Excel
    """
    cutoff = datetime.now().timestamp() - keep_days * 86400
    cleaned = 0

    for folder in [NEWS_DIR, MEMORY_DIR, EXCEL_DIR]:
        for f in folder.glob('**/*'):
            if f.is_file() and f.stat().st_mtime < cutoff:
                f.unlink()
                cleaned += 1

    print(f"[STORE] Cleanup: removed {cleaned} old files (>{keep_days} days)")
    return cleaned


def export_all_prices_excel() -> Path:
    """
    Export ALL stored price CSVs into one mega Excel file.
    One sheet per stock. Useful for full offline analysis.
    """
    path = EXCEL_DIR / f"all_prices_{today_str()}.xlsx"
    symbols = list_stored_symbols()

    if not symbols:
        print("[STORE] No price CSVs found to export")
        return None

    print(f"[STORE] Exporting {len(symbols)} stocks to Excel...")

    try:
        with pd.ExcelWriter(str(path), engine='openpyxl') as writer:
            for sym in symbols[:100]:   # Excel limit: 100 sheets
                df = load_prices_csv(sym)
                if df is not None and not df.empty:
                    # Add derived columns
                    df['Daily_Return_%'] = df['Close'].pct_change() * 100
                    df['SMA_20'] = df['Close'].rolling(20).mean()
                    df['SMA_50'] = df['Close'].rolling(50).mean()
                    df['SMA_200'] = df['Close'].rolling(200).mean()
                    df.round(2).to_excel(writer, sheet_name=sym[:31])

        print(f"[STORE] All prices exported: {path}")
        return path
    except Exception as e:
        print(f"[STORE] Export error: {e}")
        return None
